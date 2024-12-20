import openpyxl as xl 
wb =xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell =sheet.cell(1,1)

for row in range(2,sheet.max_row+1):
    cell=sheet.cell(row,3)
    new_price= cell.value * 0.9
    new_price_cell =sheet.cell(row,4)
    new_price_cell.value = '$'+ new_price
    print(new_price_cell)
#wb.save('transactions2.xlsx')


